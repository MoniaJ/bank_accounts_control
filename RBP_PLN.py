import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.styles import Color, Fill, Font, Alignment
from tkinter import Tk
from tkinter.filedialog import askopenfilename

Tk().withdraw()
filename = askopenfilename() # show an "Open" dialog box and return the path to the selected file
print(filename)

raport = Workbook()
#ws = raport.create_sheet()
ws = raport.active
ws.title = 'sheet 1'

#filename = r"C:/Users/Monika/Desktop/eoffice/RBP_PLN.xml"

headers = {1:"Company / No.of transfer", 2: "Amount", 4: "Supplier", 5: "Supplier's bank account number", 8:"Number(s) of invoice(s)"}
act_columns = [1,2,4,5,8]
for col in act_columns:
    ws.cell(column=col, row=1).value = headers[col]

ft = Font(name='Arial',size=8, italic=True)
wt = Alignment(wrap_text=True, horizontal='center', vertical='center')
for col in range (1,9):
    ws.cell(column=col, row=1).font = ft
    ws.cell(column=col, row=1).alignment = wt

    ws.freeze_panes = 'A2'
tree = ET.parse(filename)
root = tree.getroot()

for child in root:  #child is Cstmr...
    #for header in child:
        #print(header.tag[48:len(header.tag)], len(header.tag))         #prints GrpHdr 54 i PmtInf 54    
    count_row = 2
    for line in child[1]:                                               #ei. only for rows in PmtInf        
        #print(line.tag[48:len(line.tag)], len(line.tag))
        if line.tag[48:len(line.tag)] == "CdtTrfTxInf":
            #print(line.tag[48:len(line.tag)], len(line.tag))
            count_col = 1
            
            for info in line:
                try:
                    if count_col == 2:
                        amount = info[0].text.replace('.',',')
                        ws.cell(column=count_col, row=count_row).value = amount
                    else:
                        ws.cell(column=count_col, row=count_row).value = info[0].text
                except:
                    ws.cell(column=count_col, row=count_row).value = 0
                if info.tag[48:len(info.tag)] == 'CdtrAcct':    
                    ws.cell(column=count_col, row=count_row).value = info[0][0][0].text
                count_col +=1 
            count_row += 1
            
            
                    
raport.save('RBP_PLN.xlsx')
